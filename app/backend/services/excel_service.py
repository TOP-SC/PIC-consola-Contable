from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from config import app_dir
from models.schemas import ImpuestoRow
from services.ibper import parse_ibper_line


def _sample_txt_root() -> Path:
    # En desarrollo: Contable/Archivos de texto IBPER/Documentos
    # En portable: opcional junto al exe
    candidates = [
        app_dir().parent / "Archivos de texto IBPER" / "Documentos",
        app_dir() / "muestras" / "Documentos",
    ]
    for c in candidates:
        if c.exists():
            return c
    return candidates[0]


SAMPLE_TXT_ROOT = _sample_txt_root()


def rows_to_excel_bytes(rows: list[ImpuestoRow]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "IBPER"

    headers = [
        "LOCAL",
        "T_COMP",
        "N_COMP",
        "FECHA_EMIS",
        "RAZON_SOCI",
        "IDENTIFTRI",
        "IMPORTE_GRAVADO",
        "IMPORTE",
        "ALICUOTA",
        "DESC_ALIC",
        "PORCE_IMP",
        "IMPUESTO",
    ]
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(color="FFFFFF", bold=True)

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row.local,
            row.t_comp,
            row.n_comp,
            row.fecha_emis.isoformat(),
            row.razon_soci,
            row.identiftri,
            row.importe_gravado,
            row.importe,
            row.alicuota,
            row.desc_alic,
            row.porce_imp,
            row.impuesto,
        ]
        for c_idx, value in enumerate(values, start=1):
            ws.cell(r_idx, c_idx, value)

    for col in ws.columns:
        max_len = 10
        letter = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, min(len(str(cell.value)), 40))
        ws.column_dimensions[letter].width = max_len + 2

    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def load_demo_rows(desde=None, hasta=None) -> list[ImpuestoRow]:
    """Carga muestras locales de Documentos si SQL no está disponible."""
    rows: list[ImpuestoRow] = []
    if not SAMPLE_TXT_ROOT.exists():
        return rows

    for path in sorted(SAMPLE_TXT_ROOT.rglob("*.txt")):
        try:
            text = path.read_text(encoding="latin-1", errors="replace")
        except Exception:
            continue
        for line in text.splitlines():
            parsed = parse_ibper_line(line)
            if not parsed:
                continue
            if desde and parsed.fecha_emis < desde:
                continue
            if hasta and parsed.fecha_emis > hasta:
                continue
            rows.append(parsed)

    rows.sort(key=lambda r: (r.t_comp, r.n_comp, r.fecha_emis.isoformat()))
    return rows
